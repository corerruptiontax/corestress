# db.py
from openpyxl import load_workbook
import pandas as pd
from colorama import Fore, Style
from tqdm import tqdm
import os

def full_vlookup(template_file, loc_data):
    """
    Melakukan VLOOKUP data dari file KTP ke template
    
    Parameter:
    template_file (str): Path file template Excel
    loc_data (dict): Data konfigurasi lokasi
    """
    print(Fore.CYAN + "\n=== PROSES VLOOKUP ===" + Style.RESET_ALL)
    
    try:
        # Input file KTP
        ktp_file = input(Fore.GREEN + "Nama file KTP (contoh: KTP.xlsx): " + Style.RESET_ALL).strip()
        if not ktp_file.endswith('.xlsx'):
            ktp_file += '.xlsx'
            
        if not os.path.exists(ktp_file):
            raise FileNotFoundError(f"File KTP '{ktp_file}' tidak ditemukan!")

        # Validasi sheet KTP
        ktp_sheet_name = loc_data['ktp_sheet']
        if ktp_sheet_name not in pd.ExcelFile(ktp_file).sheet_names:
            raise ValueError(f"Sheet '{ktp_sheet_name}' tidak ada di file KTP!")

        # Buka template
        print(Fore.BLUE + "📂 Membuka template Excel..." + Style.RESET_ALL)
        wb = load_workbook(template_file)
        sheet = wb['Faktur']
        
        # Ambil kode BBN
        kode_bbn = [cell[0].value for cell in sheet.iter_rows(min_row=4, min_col=18, max_col=18)]
        
        # Baca data KTP
        print(Fore.BLUE + "🔍 Membaca data KTP..." + Style.RESET_ALL)
        ktp_df = pd.read_excel(
            ktp_file,
            sheet_name=ktp_sheet_name,
            usecols="C,D,E,H,I,J,K",
            header=0,
            dtype=str
        )
        
        # Proses mapping
        updated = 0
        for idx, kode in tqdm(enumerate(kode_bbn, 4), total=len(kode_bbn), desc="VLOOKUP"):
            if pd.isna(kode) or kode == "":
                continue
            
            result = ktp_df[ktp_df.iloc[:, 0] == str(kode)]
            if not result.empty:
                data = result.iloc[0]
                sheet[f'S{idx}'] = data.iloc[0]  # Kolom S: Kode Barang
                sheet[f'K{idx}'] = data.iloc[1]  # Kolom K: Jenis ID Pembeli
                sheet[f'M{idx}'] = data.iloc[2]  # Kolom M: Negara Pembeli
                sheet[f'J{idx}'] = data.iloc[3]  # Kolom J: Email Pembeli
                sheet[f'Q{idx}'] = data.iloc[4]  # Kolom Q: ID TKU Pembeli
                sheet[f'N{idx}'] = data.iloc[5]  # Kolom N: Nama Pembeli
                sheet[f'O{idx}'] = data.iloc[6]  # Kolom O: Alamat Pembeli
                updated += 1
        
        # Simpan file
        output_file = template_file.replace(".xlsx", "_FULL.xlsx")
        print(Fore.BLUE + "💾 Menyimpan hasil..." + Style.RESET_ALL)
        wb.save(output_file)
        
        print(Fore.GREEN + f"✅ Diupdate: {updated}/{len(kode_bbn)}" + Style.RESET_ALL)
        print(Fore.BLUE + f"📁 File hasil: {output_file}" + Style.RESET_ALL)
        
    except Exception as e:
        error_msg = f"""
        ❌ ERROR VLOOKUP:
        File: {os.path.basename(ktp_file) if 'ktp_file' in locals() else 'N/A'}
        Detil Error: {str(e)}
        """
        print(Fore.RED + error_msg + Style.RESET_ALL)
