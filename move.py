import pandas as pd    
from openpyxl import load_workbook    
import argparse  # Import argparse untuk menangani argumen  
  
def populate_detail_faktur(ff_file, source_file):    
    # Load workbook dari file tujuan  
    wb = load_workbook(ff_file)    
    detail_faktur_sheet = wb['DetailFaktur']    
  
    # Membaca data dari file sumber    
    try:    
        source_data = pd.read_excel(source_file)    
    except FileNotFoundError:    
        print("File tidak ditemukan. Pastikan nama file sudah benar.")    
        return    
    
    # Mengisi header    
    headers = ["Baris", "Barang/Jasa", "Kode Barang dan Jasa", "Nama Barang", "Nama Satuan Ukur",    
               "Harga Satuan", "Jumlah Barang Jasa", "Total Diskon", "DPP", "DPP Nilai Lain",    
               "Tarif PPN", "PPN", "Tarif PPnBM", "PPnBM"]    
    
    for col_num, header in enumerate(headers, 1):    
        detail_faktur_sheet.cell(row=1, column=col_num, value=header)    
    
    # Mengisi data    
    current_row = 2  # Mulai dari baris kedua    
    for index, row in source_data.iterrows():    
        if pd.notna(row['Nama Barang']) and row['Nama Barang'] != "":    
            detail_faktur_sheet.cell(row=current_row, column=1, value=row['Baris'])    
            detail_faktur_sheet.cell(row=current_row, column=2, value="A")    
            detail_faktur_sheet.cell(row=current_row, column=3, value="000000")    
            detail_faktur_sheet.cell(row=current_row, column=4, value=row['Nama Barang'])    
            detail_faktur_sheet.cell(row=current_row, column=5, value=row.get('Nama Satuan Ukur', 'UM.0002'))    
              
            # Menghitung DPP  
            harga_satuan = round(row['Harga DPP'], 2)  
            jumlah_barang = row['Qty']  
            dpp = round(harga_satuan * jumlah_barang, 2)  
            detail_faktur_sheet.cell(row=current_row, column=6, value=harga_satuan)    
            detail_faktur_sheet.cell(row=current_row, column=7, value=jumlah_barang)    
            detail_faktur_sheet.cell(row=current_row, column=8, value=0)    
            detail_faktur_sheet.cell(row=current_row, column=9, value=dpp)    
              
            # Menghitung DPP Nilai Lain  
            dpp_nilai_lain = round(dpp * (11/12), 2)  
            detail_faktur_sheet.cell(row=current_row, column=10, value=dpp_nilai_lain)    
              
            # Menghitung PPN  
            tarif_ppn = 12  # Misalkan tarif PPN adalah 12%  
            ppn = round(dpp_nilai_lain * (tarif_ppn / 100))  
            detail_faktur_sheet.cell(row=current_row, column=11, value=tarif_ppn)    
            detail_faktur_sheet.cell(row=current_row, column=12, value=ppn)    
            detail_faktur_sheet.cell(row=current_row, column=13, value=0)    
            detail_faktur_sheet.cell(row=current_row, column=14, value=0)    
  
            current_row += 1    
  
    # Menyimpan perubahan    
    wb.save(ff_file)    
    print(f"Data berhasil dipindahkan ke sheet 'DetailFaktur'. Total transaksi: {current_row - 2}")    
    
if __name__ == "__main__":    
    # Menggunakan argparse untuk menangani argumen  
    parser = argparse.ArgumentParser(description='Proses data dari file sumber ke file tujuan.')  
    parser.add_argument('source_file', type=str, help='Nama file sumber (misal: "Faktur Pajak Desember 2024.xlsx")')  
    parser.add_argument('ff_file', type=str, help='Nama file tujuan (misal: "FF.xlsx")')  
      
    args = parser.parse_args()  
      
    populate_detail_faktur(args.ff_file, args.source_file)  
