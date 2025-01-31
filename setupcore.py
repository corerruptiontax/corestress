# setupcore.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
from colorama import Fore, Style

def create_template(file_name):
    print("\n=== GENERATE TEMPLATE EXCEL ===")
    try:
        file_name = file_name.strip()  # Menghapus spasi di awal dan akhir
        file_path = f"{file_name}.xlsx"  # Tambahkan .xlsx saat menyimpan
        
        # Pilih lokasi
        print("Pilih lokasi NPWP:")
        print("1. Surabaya\n2. Semarang\n3. Samarinda\n4. Bagong Jaya")
        location = input("Masukkan nomor (1-4): ").strip()
        
        # Mapping data
        locations = {
            "1": {"name": "Surabaya", "npwp": "0947793543518000"},
            "2": {"name": "Semarang", "npwp": "0947793543518000"},
            "3": {"name": "Samarinda", "npwp": "0947793543518000"},
            "4": {"name": "Bagong Jaya", "npwp": "0712982594609000"}
        }
        
        if location not in locations:
            print("❌ Error: Pilihan lokasi tidak valid!")
            return
            
        loc_data = locations[location]
        
        # Buat workbook
        wb = Workbook()
        
        # Sheet Faktur
        faktur = wb.active
        faktur.title = "Faktur"
        faktur.merge_cells('A1:B1')
        faktur['A1'] = "Isi NPWP Penjual"
        faktur['A1'].font = Font(bold=True)
        faktur['A1'].alignment = Alignment(horizontal='center', vertical='center')
        faktur['C1'] = loc_data["npwp"]
        
        # Header Faktur
        headers = [
            "Baris", "Tanggal Faktur", "Jenis Faktur", "Kode Transaksi",
            "Keterangan Tambahan", "Dokumen Pendukung", "Referensi",
            "Cap Fasilitas", "ID TKU Penjual", "NPWP/NIK Pembeli",
            "Jenis ID Pembeli", "Negara Pembeli", "Nomor Dokumen Pembeli",
            "Nama Pembeli", "Alamat Pembeli", "Email Pembeli",
            "ID TKU Pembeli", "Kode Pelanggan"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = faktur.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        # Sheet DetailFaktur
        detail = wb.create_sheet("DetailFaktur")
        detail_headers = [
            "Baris", "Barang/Jasa", "Kode Barang", "Nama Barang",
            "Nama Satuan Ukur", "Harga Satuan", "Jumlah Barang Jasa",
            "Total Diskon", "DPP", "DPP Nilai Lain", "Tarif PPN",
            "PPN", "Tarif PPnBM", "PPnBM"
        ]
        for col_num, header in enumerate(detail_headers, 1):
            detail.cell(row=1, column=col_num, value=header)
        
        # Sheet Lokasi
        wb.create_sheet(loc_data["name"])
        
        # Adjust column widths
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value)) for cell in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
        
        wb.save(file_path)
        print(Fore.GREEN + f"\n✅ File berhasil dibuat: {os.path.abspath(file_path)}" + Style.RESET_ALL)
        
    except Exception as e:
        print(Fore.RED + f"\n❌ Error: {str(e)}" + Style.RESET_ALL)

if __name__ == "__main__":
    create_template("template")  # Contoh pemanggilan
