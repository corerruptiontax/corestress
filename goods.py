# goods.py
import pandas as pd  
from openpyxl import load_workbook  
from colorama import Fore, Style
from tqdm import tqdm
import os

def populate_detail_faktur(template_file, source_file):  
    """
    Mengisi data barang ke sheet DetailFaktur
    
    Parameter:
    template_file (str): Path file template Excel
    source_file (str): Path file sumber data barang
    """
    print(Fore.CYAN + "\n=== PROSES DATA BARANG ===" + Style.RESET_ALL)
    
    try:
        # Validasi file sumber
        if not os.path.exists(source_file):
            raise FileNotFoundError(f"File sumber '{source_file}' tidak ditemukan!")
            
        if not source_file.endswith('.xlsx'):
            source_file += '.xlsx'

        # Buka template
        print(Fore.BLUE + "📂 Membuka template Excel..." + Style.RESET_ALL)
        wb = load_workbook(template_file)  
        sheet = wb['DetailFaktur']
        
        # Baca data sumber
        print(Fore.BLUE + "🔍 Membaca file sumber..." + Style.RESET_ALL)
        source_data = pd.read_excel(source_file)
        
        # Validasi kolom wajib
        required_cols = ['Nama Barang', 'Harga DPP', 'Qty']
        missing = [col for col in required_cols if col not in source_data.columns]
        if missing:
            raise ValueError(f"Kolom wajib tidak ditemukan: {missing}")

        # Hapus data lama (jika ada)
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row)

        # Proses data
        print(Fore.BLUE + "🔄 Memproses baris data..." + Style.RESET_ALL)
        current_row = 2
        for _, row in tqdm(source_data.iterrows(), desc="Progress", unit="row"):
            if pd.notna(row.get('Nama Barang')) and row['Nama Barang'] != "":  
                # Isi data dasar
                sheet.cell(row=current_row, column=1, value=row.get('Baris', ''))  
                sheet.cell(row=current_row, column=2, value=row.get('Barang/Jasa', 'A'))  
                sheet.cell(row=current_row, column=3, value=row.get('Kode Barang', '000000'))  
                sheet.cell(row=current_row, column=4, value=row['Nama Barang'])  
                sheet.cell(row=current_row, column=5, value=row.get('Nama Satuan Ukur', 'UM.0002'))  

                # Hitung nilai numerik
                harga_satuan = round(row.get('Harga DPP', 0), 2)  
                jumlah_barang = row.get('Qty', 0)  
                dpp = round(harga_satuan * jumlah_barang, 2)  
                dpp_nilai_lain = round(dpp * (11 / 12), 2)  
                ppn = round(dpp_nilai_lain * 0.12, 0)  

                # Isi nilai kalkulasi
                sheet.cell(row=current_row, column=6, value=harga_satuan)  
                sheet.cell(row=current_row, column=7, value=jumlah_barang)  
                sheet.cell(row=current_row, column=8, value=0)  # Diskon  
                sheet.cell(row=current_row, column=9, value=dpp)  
                sheet.cell(row=current_row, column=10, value=dpp_nilai_lain)  
                sheet.cell(row=current_row, column=11, value=12)  # Tarif PPN (%)  
                sheet.cell(row=current_row, column=12, value=ppn)  
                sheet.cell(row=current_row, column=13, value=0)  # Tarif PPnBM  
                sheet.cell(row=current_row, column=14, value=0)  # PPnBM  

                current_row += 1

        # Simpan perubahan
        print(Fore.BLUE + "💾 Menyimpan perubahan..." + Style.RESET_ALL)
        wb.save(template_file)
        print(Fore.GREEN + f"✅ Berhasil diproses: {current_row-2} baris" + Style.RESET_ALL)

    except Exception as e:
        error_msg = f"""
        ❌ ERROR PROSES BARANG:
        File: {os.path.basename(source_file) if 'source_file' in locals() else 'N/A'}
        Baris Error: {current_row if 'current_row' in locals() else 'N/A'}
        Detil Error: {str(e)}
        """
        print(Fore.RED + error_msg + Style.RESET_ALL)
