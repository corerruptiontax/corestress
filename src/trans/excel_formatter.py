import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def save_to_excel(data, file_name, columns):
    # Pastikan file_name memiliki ekstensi .xlsx dengan benar
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    df_process = pd.DataFrame(data, columns=columns)

    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        df_process.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        for cell in worksheet[1]:
            cell.alignment = cell.alignment.copy(horizontal='left')

    book = load_workbook(file_name)
    sheet = book['Sheet1']

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    last_row = sheet.max_row
    last_col = len(columns)

    for row in sheet[f"A1:{chr(65 + last_col - 1)}{last_row}"]:
        for cell in row:
            cell.border = border

    book.save(file_name)