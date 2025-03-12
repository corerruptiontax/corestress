# setupcore.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
from colorama import Fore, Style
from src.utils import convert_date
from pathlib import Path

def create_template(file_name, loc_data):
    print(Fore.CYAN + "\n=== GENERATE TEMPLATE ===" + Style.RESET_ALL)
    try:
        # Tentukan folder penyimpanan template
        template_folder = Path(__file__).resolve().parent.parent / "template"

        # Buat folder jika belum ada
        if not os.path.exists(template_folder):
            os.makedirs(template_folder)

        # Path lengkap untuk menyimpan template
        file_path = os.path.join(template_folder, file_name)  # Jangan tambahkan .xlsx lagi

        wb = Workbook()
        faktur = wb.active
        faktur.title = "Faktur"

        # Header NPWP
        faktur.merge_cells('A1:B1')
        faktur['A1'] = "Isi NPWP Penjual"
        faktur['A1'].font = Font(bold=True)
        faktur['A1'].alignment = Alignment(horizontal='center', vertical='center')
        faktur['C1'] = loc_data['npwp']

        # Header Faktur
        headers = [
            "Baris", 
            "Tanggal Faktur", 
            "Jenis Faktur", 
            "Kode Transaksi",
            "Keterangan Tambahan", 
            "Dokumen Pendukung", 
            "Period Dok Pendukung",
            "Referensi", 
            "Cap Fasilitas", 
            "ID TKU Penjual", 
            "NPWP/NIK Pembeli",
            "Jenis ID Pembeli", 
            "Negara Pembeli", 
            "Nomor Dokumen Pembeli",
            "Nama Pembeli", 
            "Alamat Pembeli", 
            "Email Pembeli",
            "ID TKU Pembeli", 
            "Kode Pelanggan"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = faktur.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Sheet DetailFaktur
        detail = wb.create_sheet("DetailFaktur")
        detail_headers = [
            "Baris",
            "Barang/Jasa", 
            "Kode Barang Jasa", 
            "Nama Barang/Jasa",
            "Nama Satuan Ukur", 
            "Harga Satuan", 
            "Jumlah Barang Jasa",
            "Total Diskon", 
            "DPP", 
            "DPP Nilai Lain", 
            "Tarif PPN",
            "PPN", 
            "Tarif PPnBM", 
            "PPnBM"
        ]
        for col_num, header in enumerate(detail_headers, 1):
            detail.cell(row=1, column=col_num, value=header)

        # Sheet Lokasi
        wb.create_sheet(loc_data['name'])

        # Adjust column widths
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value)) for cell in col if cell.value is not None)
                adjusted_width = (max_len + 2) * 1.2 if max_len else 10
                sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        # Simpan workbook ke folder yang sudah ditentukan
        wb.save(file_path)
        print(Fore.GREEN + f"✅ Template berhasil disimpan di: {os.path.abspath(file_path)}" + Style.RESET_ALL)

    except Exception as e:
        print(Fore.RED + f"❌ Error: {str(e)}" + Style.RESET_ALL)

if __name__ == "__main__":
    create_template("template", {'name':'Test','npwp':'123','id_tku':'456'})
