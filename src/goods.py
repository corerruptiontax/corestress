# goods.py
import pandas as pd  
from openpyxl import load_workbook  
from colorama import Fore, Style
from tqdm import tqdm
import os
from config.mappings import BARANG_MAPPING

def populate_detail_faktur(template_file, source_file):  
    print(Fore.CYAN + "\n=== PROSES DATA BARANG ===" + Style.RESET_ALL)
    
    try:
        # Validasi file
        if not os.path.exists(source_file):
            raise FileNotFoundError(f"File sumber '{source_file}' tidak ditemukan!")
            
        if not source_file.endswith('.xlsx'):
            source_file += '.xlsx'

        # Buka template
        wb = load_workbook(template_file)  
        sheet = wb['DetailFaktur']
        
        # Baca data sumber
        source_data = pd.read_excel(source_file)
        
        # Validasi kolom
        required_cols = ['Nama Barang', 'Harga DPP', 'Qty']
        missing = [col for col in required_cols if col not in source_data.columns]
        if missing:
            raise ValueError(f"Kolom wajib tidak ditemukan: {missing}")

        # Hapus data lama
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row)

        # Proses data
        current_row = 2
        for idx, row in tqdm(source_data.iterrows(), desc="Memproses", unit="row"):
            try:
                nama_barang = str(row['Nama Barang'])
                
                # Menangani baris kosong
                if pd.isna(row['Nama Barang']) and pd.isna(row['Harga DPP']) and pd.isna(row['Qty']):
                    continue  # Melewati baris jika semua kolom kosong
                
                # Handle nilai numeric
                harga_dpp = pd.to_numeric(row['Harga DPP'], errors='coerce')
                qty = pd.to_numeric(row['Qty'], errors='coerce')
                
                if pd.isna(harga_dpp) or pd.isna(qty):
                    print(Fore.YELLOW + f"⚠️ Baris {idx+2}: Nilai Harga/Qty invalid. Baris dilewati." + Style.RESET_ALL)
                    continue
                    
                # Mapping barang
                kode_barang = '000000'
                satuan_ukur = 'UM.0002'
                for keyword, (kode, satuan) in BARANG_MAPPING.items():
                    if keyword in nama_barang:
                        kode_barang = kode
                        satuan_ukur = satuan
                        break
                
                # Hitung nilai
                harga_satuan = round(float(harga_dpp), 2)
                jumlah_barang = int(qty)
                dpp = round(harga_satuan * jumlah_barang, 2)
                dpp_nilai_lain = round(dpp * (11 / 12), 2)
                ppn = round(dpp_nilai_lain * 0.12, 0)  # 2 desimal

                # Isi data
                sheet.cell(row=current_row, column=1, value=row.get('Baris', ''))
                sheet.cell(row=current_row, column=2, value='A')
                sheet.cell(row=current_row, column=3, value=kode_barang)
                sheet.cell(row=current_row, column=4, value=nama_barang)
                sheet.cell(row=current_row, column=5, value=satuan_ukur)
                sheet.cell(row=current_row, column=6, value=harga_satuan)
                sheet.cell(row=current_row, column=7, value=jumlah_barang)
                sheet.cell(row=current_row, column=8, value=0)
                sheet.cell(row=current_row, column=9, value=dpp)
                sheet.cell(row=current_row, column=10, value=dpp_nilai_lain)
                sheet.cell(row=current_row, column=11, value=12)
                sheet.cell(row=current_row, column=12, value=ppn)
                sheet.cell(row=current_row, column=13, value=0)
                sheet.cell(row=current_row, column=14, value=0)

                current_row += 1
                
            except Exception as e:
                print(Fore.RED + f"❌ Error Baris {idx+2}: {str(e)}" + Style.RESET_ALL)
                continue

        # Tambah END
        sheet.cell(row=current_row, column=1, value="END")
        
        wb.save(template_file)
        print(Fore.GREEN + f"✅ Berhasil diproses: {current_row-2} baris" + Style.RESET_ALL)

    except Exception as e:
        print(Fore.RED + f"❌ Error: {str(e)}" + Style.RESET_ALL)
