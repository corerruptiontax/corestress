# db.py
from openpyxl import load_workbook
import pandas as pd
from colorama import Fore, Style
from tqdm import tqdm
import os
import datetime
from pathlib import Path

def full_vlookup(template_file, loc_data):
    print(Fore.CYAN + "\n=== PROSES VLOOKUP ===" + Style.RESET_ALL)
    try:
        # Default filename
        DEFAULT_KTP_FILE = "KTP - list (NEW).xlsx"
        
        # Input file KTP dengan opsi default
        ktp_file = input(
            Fore.GREEN + f"Nama file DEFAULT : [{DEFAULT_KTP_FILE}]: " + 
            Style.RESET_ALL
        ).strip()

        # Gunakan default jika input kosong
        if not ktp_file:
            ktp_file = DEFAULT_KTP_FILE
        
        # Tambahkan ekstensi .xlsx jika belum ada
        if not ktp_file.endswith('.xlsx'):
            ktp_file += '.xlsx'

        # Validasi file sampai benar-benar ada
        while not os.path.exists(ktp_file):
            print(Fore.RED + f"❌ File '{ktp_file}' tidak ditemukan!" + Style.RESET_ALL)
            ktp_file = input(Fore.GREEN + "Masukkan ulang nama file KTP: " + Style.RESET_ALL).strip()
            if not ktp_file.endswith('.xlsx'):
                ktp_file += '.xlsx'

        # Buka template
        print(Fore.BLUE + "📂 Membuka template Excel..." + Style.RESET_ALL)
        wb = load_workbook(template_file)
        sheet = wb['Faktur']
        
        # Ambil kode BBN (kolom R) dan skip baris END
        kode_bbn = []
        for row in sheet.iter_rows(min_row=4, min_col=19, max_col=19):  # Kolom R = 18
            cell_value = row[0].value
            if cell_value is not None and str(cell_value).strip().upper() != "END":
                kode_bbn.append(cell_value)
        
        # Baca data KTP
        print(Fore.BLUE + "🔍 Membaca data KTP..." + Style.RESET_ALL)
        ktp_df = pd.read_excel(
            ktp_file,
            sheet_name=loc_data['ktp_sheet'],
            usecols="C,D,E,H,I,J,K",
            header=0,
            dtype=str
        )
        
        # Proses mapping
        updated = 0
        total_valid = len(kode_bbn)
        
        for idx, kode in tqdm(enumerate(kode_bbn, 4), total=total_valid, desc="VLOOKUP"):
            result = ktp_df[ktp_df.iloc[:, 0] == str(kode)]
            if not result.empty:
                data = result.iloc[0]
                sheet[f'T{idx}'] = data.iloc[0]  # Kode Pelanggan (sekarang di kolom T)
                sheet[f'L{idx}'] = data.iloc[1]  # NPWP/NIK Pembeli (geser ke L)
                sheet[f'N{idx}'] = data.iloc[2]  # Negara Pembeli (geser ke N)
                sheet[f'K{idx}'] = data.iloc[3]  # ID TKU Penjual (geser ke K)
                sheet[f'R{idx}'] = data.iloc[4]  # Email Pembeli (geser ke R)
                sheet[f'O{idx}'] = data.iloc[5]  # Nama Pembeli (geser ke O)
                sheet[f'P{idx}'] = data.iloc[6]  # Alamat Pembeli (geser ke P)
                updated += 1
        
        # Buat folder 'imp' jika belum ada
        imp_folder = Path(__file__).resolve().parent.parent / "imp"
        
        # Buat folder jika belum ada
        if not imp_folder.exists():
            imp_folder.mkdir(parents=True)

        # Ambil nama file tanpa ekstensi
        base_name = Path(template_file).stem

        # Buat nama file unik dengan menambahkan timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file_name = f"{base_name}_FULL_{timestamp}.xlsx"

        # Gabungkan path folder dengan nama file unik
        output_file = imp_folder / output_file_name

        print(Fore.BLUE + "💾 Menyimpan hasil..." + Style.RESET_ALL)
        wb.save(output_file)

        print(Fore.GREEN + f"✅ Diupdate: {updated}/{total_valid}" + Style.RESET_ALL)
        print(Fore.BLUE + f"📁 File hasil: {output_file}" + Style.RESET_ALL)
        
    except Exception as e:
        error_msg = f"""
        ❌ ERROR VLOOKUP:
        File: {os.path.basename(ktp_file) if 'ktp_file' in locals() else 'N/A'}
        Detil Error: {str(e)}
        """
        print(Fore.RED + error_msg + Style.RESET_ALL)