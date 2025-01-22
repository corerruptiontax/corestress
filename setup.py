import pandas as pd  
from openpyxl import Workbook  
from openpyxl.styles import Font  
from openpyxl.utils import get_column_letter  
from openpyxl.styles import Alignment  
  
def create_excel_file():  
    # Meminta input nama file output tanpa ekstensi .xlsx  
    file_name = input("Masukkan nama file output (tanpa .xlsx): ")  
    file_path = f"{file_name}.xlsx"  
  
    # Menampilkan pilihan lokasi dengan angka  
    print("Pilih lokasi NPWP:")  
    print("1. Surabaya")  
    print("2. Semarang")  
    print("3. Samarinda")  
    print("4. Bagong Jaya")  
  
    # Meminta input lokasi berupa angka  
    location_choice = input("Masukkan nomor lokasi (1, 2, 3, atau 4): ").strip()  
  
    # Validasi input lokasi  
    if location_choice not in ["1", "2", "3", "4"]:  
        print("Pilihan lokasi tidak valid. Harap masukkan 1, 2, 3, atau 4.")  
        return  
  
    # Mapping pilihan angka ke lokasi dan NPWP/ID TKU  
    location_data = {  
        "1": {"location": "Surabaya", "npwp": "0947793543518000", "id_tku": "0947793543518000000000"},  
        "2": {"location": "Semarang", "npwp": "0947793543518000", "id_tku": "0947793543518000000001"},  
        "3": {"location": "Samarinda", "npwp": "0947793543518000", "id_tku": "0947793543518000000002"},  
        "4": {"location": "Bagong Jaya", "npwp": "0712982594609000", "id_tku": "0712982594609000000000"}  
    }  
  
    location_info = location_data[location_choice]  
  
    # Membuat workbook dan sheet  
    wb = Workbook()  
    faktur_sheet = wb.active  
    faktur_sheet.title = "Faktur"  
    detail_faktur_sheet = wb.create_sheet(title="DetailFaktur")  
      
    # Menambahkan sheet baru dengan nama lokasi yang dipilih  
    location_sheet = wb.create_sheet(title=location_info["location"])  
  
    # Mengisi sheet "Faktur"  
    faktur_sheet.merge_cells('A1:B1')  
    faktur_sheet['A1'] = "Isi NPWP Penjual"  
    faktur_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')  
    faktur_sheet['A1'].font = Font(bold=True)  # Membuat A1 bold  
    faktur_sheet['C1'] = location_info["npwp"]  
  
    faktur_headers = [  
        "Baris", "Tanggal Faktur", "Jenis Faktur", "Kode Transaksi", "Keterangan Tambahan",  
        "Dokumen Pendukung", "Refrensi", "Cap Fasilitas", "ID TKU Penjual",  
        "NPWP/NIK Pembeli", "Jenis ID Pembeli", "Negara Pembeli", "Nomor Dokumen Pembeli",  
        "Nama Pembeli", "Alamat Pembeli", "Email Pembeli", "ID TKU Pembeli"  
    ]  
  
    for col_num, header in enumerate(faktur_headers, 1):  
        cell = faktur_sheet.cell(row=3, column=col_num)  
        cell.value = header  
        cell.font = Font(bold=True)  
  
    # Menampilkan "ID TKU Penjual" di I3 dan ID TKU di I4  
    faktur_sheet['I3'] = "ID TKU Penjual"  
    faktur_sheet['I4'] = location_info["id_tku"]  
  
    # Mengisi sheet "DetailFaktur"  
    detail_faktur_headers = [  
        "Baris", "Barang/Jasa", "Kode Barang Jasa", "Nama Barang/Jasa", "Nama Satuan Ukur",  
        "Harga Satuan", "Jumlah Barang Jasa", "Total Diskon", "DPP", "DPP Nilai Lain",  
        "Tarif PPN", "PPN", "Tarif PPnBM", "PPnBM"  
    ]  
  
    for col_num, header in enumerate(detail_faktur_headers, 1):  
        cell = detail_faktur_sheet.cell(row=1, column=col_num)  
        cell.value = header  
  
    # Menyesuaikan lebar kolom  
    for sheet in wb.worksheets:  
        for column_cells in sheet.columns:  
            length = max(len(str(cell.value)) for cell in column_cells)  
            adjusted_length = (length + 2)  
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_length  
  
    # Menyimpan file Excel  
    wb.save(file_path)  
    print(f"File Excel '{file_path}' berhasil dibuat.")  
  
if __name__ == "__main__":  
    create_excel_file()  
